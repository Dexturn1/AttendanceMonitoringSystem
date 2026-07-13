############################################# IMPORTING ################################################
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mess
from tkinter import PhotoImage
from PIL import Image, ImageTk
import tkinter.simpledialog as tsd
import cv2
import os
import csv
import numpy as np
from PIL import Image
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime
import time
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import sys
import platform  # For Caps Lock detection
from subprocess import check_output, CalledProcessError

############################################# FUNCTIONS ################################################

def assure_path_exists(path):
    os.makedirs(path, exist_ok=True)

##################################################################################

def tick():
    # Get the current time
    current_time = time.strftime('%I:%M:%S %p')
    # Update the clock label with the current time
    clock.config(text=current_time)
    # Schedule the next update after 1000 milliseconds (1 second)
    clock.after(1000, tick)

###################################################################################

def contact():
    mess._show(title='Contact us', message="Please contact us on : 'dasdarshan7@gmail.com' ")

###################################################################################

def check_haarcascadefile():
    exists = os.path.isfile("haarcascade_frontalface_default.xml")
    if exists:
        return True
    else:
        mess._show(title='Some file missing', message='Please make sure "haarcascade_frontalface_default.xml" exists in the application folder.')
        return False

######################################## CAPS LOCK DETECTION ########################################

def is_capslock_on():
    """
    Cross-platform Caps Lock state checker.
    Works on Windows, macOS, and common Linux setups.
    Returns True if Caps Lock is ON, False otherwise.
    """
    system = platform.system()
    try:
        # --- Windows ---
        if system == "Windows":
            import ctypes
            hll = ctypes.WinDLL("User32.dll")
            VK_CAPITAL = 0x14
            state = hll.GetKeyState(VK_CAPITAL)
            return bool(state & 0x0001)  # low-order bit indicates toggle
        # --- macOS ---
        elif system == "Darwin":
            try:
                from Quartz import CGEventSourceKeyState, kCGEventSourceStateHIDSystemState
                # 57 is the Caps Lock virtual key on macOS
                return bool(CGEventSourceKeyState(kCGEventSourceStateHIDSystemState, 57))
            except Exception:
                # pyobjc may not be installed — can't detect
                return False
        # --- Linux ---
        elif system == "Linux":
            # Try xset (X11)
            try:
                out = check_output(["xset", "q"]).decode(errors='ignore')
                return "Caps Lock:   on" in out or "Caps Lock: on" in out
            except (CalledProcessError, FileNotFoundError):
                # xset not available (Wayland or headless) — no reliable detection
                return False
        else:
            return False
    except Exception:
        return False

# Continuous Caps Lock warning for any Entry field
def check_capslock_continuous(entry_widget, warning_label):
    try:
        if is_capslock_on():
            warning_label.config(text="⚠ Caps Lock is ON!")
        else:
            warning_label.config(text="")
    except Exception:
        warning_label.config(text="")
    entry_widget.after(200, lambda: check_capslock_continuous(entry_widget, warning_label))

###################################################################################

def autofit_excel(file_path):
    """
    Auto-fit column widths for all sheets in the given workbook file_path (openpyxl).
    """
    try:
        wb = load_workbook(file_path)
    except Exception:
        return
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            try:
                col_letter = get_column_letter(col[0].column)
            except Exception:
                continue
            for cell in col:
                try:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width < 8:
                adjusted_width = 8
            ws.column_dimensions[col_letter].width = adjusted_width
    try:
        wb.save(file_path)
    except Exception:
        pass

###################################################################################

def save_pass():
    assure_path_exists("TrainingImageLabel/")
    exists1 = os.path.isfile(os.path.join("TrainingImageLabel", "psd.txt"))
    if exists1:
        tf = open(os.path.join("TrainingImageLabel", "psd.txt"), "r")
        key = tf.read()
        tf.close()
    else:
        master.destroy()
        new_pas = tsd.askstring('Old Password not found', 'Please enter a new password below', show='*')
        if new_pas is None:
            mess._show(title='No Password Entered', message='Password not set!! Please try again')
            return
        else:
            with open(os.path.join("TrainingImageLabel", "psd.txt"), "w") as tf:
                tf.write(new_pas)
            mess._show(title='Password Registered', message='New password was registered successfully!!')
            return
    op = (old.get())
    newp = (new.get())
    nnewp = (nnew.get())
    if (op == key):
        if (newp == nnewp):
            with open(os.path.join("TrainingImageLabel", "psd.txt"), "w") as txf:
                txf.write(newp)
        else:
            mess._show(title='Error', message='Confirm new password again!!!')
            return
    else:
        mess._show(title='Wrong Password', message='Please enter correct old password.')
        return
    mess._show(title='Password Changed', message='Password changed successfully!!')
    master.destroy()

###################################################################################

def change_pass():
    global master
    master = tk.Toplevel(window)
    master.title("Change Password")
    master.geometry("450x280")
    master.resizable(False, False)
    master.configure(bg="#2d3436")
    master.transient(window)
    master.grab_set()

    main_frame = ttk.Frame(master, padding=20)
    main_frame.pack(fill="both", expand=True)

    ttk.Label(main_frame, text="Enter Old Password", style="TLabel").grid(row=0, column=0, sticky="w", pady=(0, 5))
    global old
    old = ttk.Entry(main_frame, width=30, show='*', font=('Helvetica', 12))
    old.grid(row=1, column=0, columnspan=2, sticky="ew")
    old_warning = ttk.Label(main_frame, text="", style="Warning.TLabel")
    old_warning.grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 10))
    old.bind("<KeyPress>", lambda e: check_capslock_continuous(old, old_warning))

    ttk.Label(main_frame, text="Enter New Password", style="TLabel").grid(row=3, column=0, sticky="w", pady=(0, 5))
    global new
    new = ttk.Entry(main_frame, width=30, show='*', font=('Helvetica', 12))
    new.grid(row=4, column=0, columnspan=2, sticky="ew")
    new_warning = ttk.Label(main_frame, text="", style="Warning.TLabel")
    new_warning.grid(row=5, column=0, columnspan=2, sticky="w", pady=(0, 10))
    new.bind("<KeyPress>", lambda e: check_capslock_continuous(new, new_warning))

    ttk.Label(main_frame, text="Confirm New Password", style="TLabel").grid(row=6, column=0, sticky="w", pady=(0, 5))
    global nnew
    nnew = ttk.Entry(main_frame, width=30, show='*', font=('Helvetica', 12))
    nnew.grid(row=7, column=0, columnspan=2, sticky="ew")
    nnew_warning = ttk.Label(main_frame, text="", style="Warning.TLabel")
    nnew_warning.grid(row=8, column=0, columnspan=2, sticky="w", pady=(0, 15))
    nnew.bind("<KeyPress>", lambda e: check_capslock_continuous(nnew, nnew_warning))

    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=9, column=0, columnspan=2, sticky="ew")
    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)

    save_button = ttk.Button(button_frame, text="Save", command=save_pass, style="Accent.TButton")
    save_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))
    cancel_button = ttk.Button(button_frame, text="Cancel", command=master.destroy)
    cancel_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))

#####################################################################################

# -- New helper: custom password dialog with visibility toggle and caps-lock warning --
def ask_password_dialog(title="Password", prompt="Enter Password", require_confirm=False):
    """
    Opens a modal Toplevel dialog that asks for a password.
    If require_confirm is True, asks for password + confirm password and returns the password only if they match.
    Returns the entered password (string) or None if cancelled.
    """
    result = {"password": None}
    dialog = tk.Toplevel(window)
    dialog.transient(window)
    dialog.grab_set()
    dialog.title(title)
    dialog.geometry("420x240" if require_confirm else "420x180")
    dialog.resizable(False, False)
    dialog.configure(bg="#2d3436")

    main_frame = ttk.Frame(dialog, padding=20)
    main_frame.pack(fill="both", expand=True)
    main_frame.columnconfigure(0, weight=1)

    ttk.Label(main_frame, text=prompt, style="TLabel").grid(row=0, column=0, sticky="w", pady=(0, 5))
    
    pwd_frame = ttk.Frame(main_frame)
    pwd_frame.grid(row=1, column=0, sticky="ew")
    pwd_frame.columnconfigure(0, weight=1)

    pwd_var = tk.StringVar()
    pwd_entry = ttk.Entry(pwd_frame, textvariable=pwd_var, width=30, show='*', font=('Helvetica', 12))
    pwd_entry.grid(row=0, column=0, sticky="ew")
    pwd_entry.focus_set()

    def toggle_pwd():
        if pwd_entry.cget('show') == '':
            pwd_entry.config(show='*')
            toggle_btn.config(text='Show')
        else:
            pwd_entry.config(show='')
            toggle_btn.config(text='Hide')

    toggle_btn = ttk.Button(pwd_frame, text='Show', command=toggle_pwd, width=6)
    toggle_btn.grid(row=0, column=1, padx=(5, 0))

    pwd_warning = ttk.Label(main_frame, text="", style="Warning.TLabel")
    pwd_warning.grid(row=2, column=0, sticky="w", pady=(0, 10))
    pwd_entry.bind("<KeyPress>", lambda e: check_capslock_continuous(pwd_entry, pwd_warning))

    if require_confirm:
        ttk.Label(main_frame, text='Confirm Password', style="TLabel").grid(row=3, column=0, sticky="w", pady=(0, 5))
        
        confirm_frame = ttk.Frame(main_frame)
        confirm_frame.grid(row=4, column=0, sticky="ew")
        confirm_frame.columnconfigure(0, weight=1)

        confirm_var = tk.StringVar()
        confirm_entry = ttk.Entry(confirm_frame, textvariable=confirm_var, width=30, show='*', font=('Helvetica', 12))
        confirm_entry.grid(row=0, column=0, sticky="ew")
        
        confirm_warning = ttk.Label(main_frame, text="", style="Warning.TLabel")
        confirm_warning.grid(row=5, column=0, sticky="w", pady=(0, 10))
        confirm_entry.bind("<KeyPress>", lambda e: check_capslock_continuous(confirm_entry, confirm_warning))

        def toggle_confirm():
            if confirm_entry.cget('show') == '':
                confirm_entry.config(show='*')
                toggle_confirm_btn.config(text='Hide')
            else:
                confirm_entry.config(show='')
                toggle_confirm_btn.config(text='Hide')

        toggle_confirm_btn = ttk.Button(confirm_frame, text='Show', command=toggle_confirm, width=6)
        toggle_confirm_btn.grid(row=0, column=1, padx=(5, 0))

    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=6, column=0, sticky="se", pady=(15, 0))
    main_frame.rowconfigure(6, weight=1)

    def on_ok():
        pwd = pwd_var.get()
        if require_confirm:
            conf = confirm_var.get()
            if pwd == "":
                mess._show(title='No Password', message='Please enter a password.')
                return
            if pwd != conf:
                mess._show(title='Mismatch', message='Passwords do not match.')
                return
        result['password'] = pwd
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    ok_btn = ttk.Button(button_frame, text="OK", width=10, command=on_ok, style="Accent.TButton")
    ok_btn.pack(side="left", padx=(0, 5))
    cancel_btn = ttk.Button(button_frame, text="Cancel", width=10, command=on_cancel)
    cancel_btn.pack(side="left")

    dialog.wait_window()
    return result['password']

###################################################################################

def psw():
    """
    Modified psw(): uses a custom dialog with show/hide toggle instead of simpledialog.askstring.
    Behavior:
    - If psd.txt exists: ask for password and if matches, call TrainImages()
    - If psd.txt does NOT exist: ask user to create a new password (with confirmation)
    """
    assure_path_exists("TrainingImageLabel/")
    psd_path = os.path.join("TrainingImageLabel", "psd.txt")
    if os.path.isfile(psd_path):
        # Ask for existing password (single entry with toggle)
        entered = ask_password_dialog(title='Password', prompt='Enter Password', require_confirm=False)
        if entered is None:
            # cancelled
            return
        with open(psd_path, "r") as tf:
            key = tf.read()
        if entered == key:
            TrainImages()
        else:
            mess._show(title='Wrong Password', message='You have entered wrong password')
    else:
        # No password exists -> ask user to create new password (confirm required)
        new_pas = ask_password_dialog(title='Set New Password', prompt='Enter new password', require_confirm=True)
        if new_pas is None:
            mess._show(title='No Password Entered', message='Password not set!! Please try again')
            return
        with open(psd_path, "w") as tf:
            tf.write(new_pas)
        mess._show(title='Password Registered', message='New password was registered successfully!!')
        return

######################################################################################

def clear():
    txt.delete(0, 'end')
    res = "1)Take Images  >>>  2)Save Profile"
    message1.configure(text=res)

def clear2():
    txt2.delete(0, 'end')
    res = "1)Take Images  >>>  2)Save Profile"
    message1.configure(text=res)

#######################################################################################

# File path for StudentDetails
student_details_folder = "StudentDetails"
student_details_file = os.path.join(student_details_folder, "StudentDetails.xlsx")

assure_path_exists(student_details_folder)

def compute_registration_count():
    try:
        df = pd.read_excel(student_details_file)
        df = df.dropna(how='all')
        if 'SERIAL NO.' in df.columns:
            valid = df['SERIAL NO.'].apply(lambda x: str(x).isdigit() if pd.notna(x) else False)
            return int(valid.sum())
        else:
            # fallback
            return max(0, len(df) - 1)
    except Exception:
        return 0

def update_registration_counter():
    """
    Updates the registration counter based on StudentDetails.xlsx accurately.
    Skips headers and empty rows.
    """
    count = 0
    if os.path.exists(student_details_file):
        try:
            df = pd.read_excel(student_details_file)
            df = df.dropna(how='all')
            if 'SERIAL NO.' in df.columns:
                valid = df['SERIAL NO.'].apply(lambda x: str(x).isdigit() if pd.notna(x) else False)
                count = int(valid.sum())
            else:
                count = max(0, len(df) - 1)
        except Exception as e:
            print("Error counting registrations:", e)
            count = 0
    message.configure(text='Total Registrations: ' + str(count))

#######################################################################################

def TakeImages():
    if not check_haarcascadefile():
        return
    assure_path_exists("StudentDetails")
    assure_path_exists("TrainingImage")

    # Ensure StudentDetails workbook exists with clean header
    if not os.path.isfile(student_details_file):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Students"
            sheet.append(['SERIAL NO.', '', 'ID', '', 'NAME'])
            workbook.save(student_details_file)
            autofit_excel(student_details_file)
        except Exception as e:
            mess._show(title='Error', message=f'Failed to create StudentDetails file: {e}')
            return

    # Load workbook to determine next serial (robustly) for registration serial
    try:
        df_existing = pd.read_excel(student_details_file)
        df_existing = df_existing.dropna(how='all')
        if 'SERIAL NO.' in df_existing.columns:
            numeric_serials = df_existing['SERIAL NO.'].apply(lambda x: int(x) if (pd.notna(x) and str(x).isdigit()) else None).dropna().astype(int)
            if len(numeric_serials) == 0:
                serial = 1
            else:
                serial = int(numeric_serials.max()) + 1
        else:
            serial = 1
    except Exception:
        # fallback
        try:
            wb_tmp = load_workbook(student_details_file)
            sheet_tmp = wb_tmp.active
            serial = sheet_tmp.max_row
            if serial <= 1:
                serial = 1
        except Exception:
            serial = 1

    Id = txt.get().strip()
    name = txt2.get().strip()

    if not Id.isdigit():
        message.configure(text="ID must be numeric.")
        return
    if not name.replace(" ", "").isalpha():
        message.configure(text="Enter a valid name (letters and spaces only).")
        return

    # Create user folder for training images
    safe_name = name.replace(" ", "_")
    user_folder = os.path.join("TrainingImage", f"{safe_name}_{Id}")
    assure_path_exists(user_folder)

    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        message.configure(text="Failed to access camera.")
        return

    detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    sampleNum = 0

    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                sampleNum += 1
                face_color = img[y:y + h, x:x + w]
                filename = f"{name}.{serial}.{Id}.{sampleNum}.jpg"
                cv2.imwrite(os.path.join(user_folder, filename), face_color)
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.imshow('Taking Images (Press Q to quit)', img)

            key = cv2.waitKey(100) & 0xFF
            if key == ord('q') or key == ord('Q') or sampleNum >= 100:
                break
    finally:
        cam.release()
        cv2.destroyAllWindows()

    # Append student details to Excel file using pandas
    try:
        df_row = pd.DataFrame([[serial, int(Id), name]], columns=['SERIAL NO.', 'ID', 'NAME'])
        try:
            df_all = pd.read_excel(student_details_file)
            df_all = df_all.dropna(how='all')
            df_all = pd.concat([df_all, df_row], ignore_index=True)
        except Exception:
            df_all = df_row
        df_all.to_excel(student_details_file, index=False)
        autofit_excel(student_details_file)
    except Exception as e:
        mess._show(title='Error', message=f'Failed to save student details: {e}')
        return

    # ✅ Update counter
    update_registration_counter()
    message1.configure(text="Images Taken! Now click Save Profile.")

########################################################################################

def TrainImages():
    if not check_haarcascadefile():
        return
    assure_path_exists("TrainingImageLabel")

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except AttributeError:
        message.configure(text="LBPH Recognizer not available. Please install opencv-contrib-python.")
        return

    faces, ID = getImagesAndLabels("TrainingImage")

    if len(faces) == 0 or len(ID) == 0:
        mess._show(title='No Registrations', message='Please register someone first!')
        return

    try:
        recognizer.train(faces, np.array(ID))
        recognizer.save(os.path.join("TrainingImageLabel", "Trainer.yml"))
        message1.configure(text="Profile trained and saved successfully.")
    except Exception as e:
        message.configure(text=f"Training failed: {e}")
        return

    # ✅ Update counter
    update_registration_counter()

############################################################################################3

def getImagesAndLabels(path):
    faces = []
    Ids = []
    # Use os.walk to go through all subdirectories
    for root, dirs, files in os.walk(path):
        for file in files:
            if file.lower().endswith(("jpeg", "jpg", "png")):
                imagePath = os.path.join(root, file)
                # Loading the image and converting it to gray scale
                try:
                    pilImage = Image.open(imagePath).convert('L')
                    imageNp = np.array(pilImage, 'uint8')
                except Exception:
                    continue
                # getting the Id from the image filename
                try:
                    ID = int(os.path.split(imagePath)[-1].split(".")[2])
                except Exception:
                    continue
                faces.append(imageNp)
                Ids.append(ID)
    return faces, Ids

###########################################################################################

def TrackImages():
    if not check_haarcascadefile():
        return
    assure_path_exists("Attendance/")
    assure_path_exists("StudentDetails/")

    # Clear Treeview
    for k in tv.get_children():
        tv.delete(k)

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except Exception:
        mess._show(title='Error', message='LBPH recognizer not available. Please install opencv-contrib-python.')
        return

    trainer_path = os.path.join("TrainingImageLabel", "Trainer.yml")
    if not os.path.isfile(trainer_path):
        mess._show(title='Data Missing', message='Please click on Save Profile to reset data!!')
        return

    recognizer.read(trainer_path)
    faceCascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    cam = cv2.VideoCapture(0)
    font = cv2.FONT_HERSHEY_SIMPLEX

    details_path = student_details_file
    if not os.path.isfile(details_path):
        mess._show(title='Details Missing', message='Students details are missing, please check!')
        cam.release()
        cv2.destroyAllWindows()
        return

    try:
        df = pd.read_excel(details_path)
    except Exception:
        df = pd.DataFrame(columns=['SERIAL NO.', 'ID', 'NAME'])

    start_time = time.time()
    max_duration = 5  # seconds
    attendance = None

    while True:
        ret, im = cam.read()
        if not ret:
            break
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = faceCascade.detectMultiScale(gray, 1.2, 5)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (225, 0, 0), 2)
            try:
                serial_pred, conf = recognizer.predict(gray[y:y + h, x:x + w])
            except Exception:
                serial_pred, conf = -1, 100

            if conf < 50:
                ts = time.time()
                date_str = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
                timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%I:%M:%S %p')
                try:
                    student_row = df.loc[df['SERIAL NO.'] == serial_pred]
                    serial_no = str(student_row['SERIAL NO.'].values[0])
                    student_id = str(student_row['ID'].values[0])
                    student_name = str(student_row['NAME'].values[0])
                    attendance = [serial_no, student_id, student_name, str(date_str), str(timeStamp)]
                    bb = student_name
                except Exception:
                    attendance = None
                    bb = "Unknown"
            else:
                bb = 'Unknown'
            cv2.putText(im, str(bb), (x, y + h), font, 1, (255, 255, 255), 2)

        cv2.imshow('Taking Attendance (Press Q to quit)', im)

        key = cv2.waitKey(1) & 0xFF
        if key in [ord('q'), ord('Q')]:
            break

        if time.time() - start_time >= max_duration:
            break

    # Save attendance to a single persistent Attendance.xlsx (never reset)
    ts = time.time()
    date_str = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
    attendance_folder = "Attendance"
    attendance_file = os.path.join(attendance_folder, "Attendance.xlsx")
    assure_path_exists(attendance_folder)

    if attendance is None:
        mess._show(title='No Entry', message='No valid attendance was captured.')
    else:
        # Create or open the workbook, ensure header
        if os.path.isfile(attendance_file):
            try:
                wb = load_workbook(attendance_file)
                ws = wb.active
            except Exception:
                # if corrupt or unreadable, recreate
                wb = Workbook()
                ws = wb.active
                ws.append(['S.No', 'ID', 'Name', 'Date', 'Time'])
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['S.No', 'ID', 'Name', 'Date', 'Time'])

        # Determine next serial number (persistent across sessions)
        try:
            if ws.max_row > 1:
                last_serial = ws.cell(row=ws.max_row, column=1).value
                try:
                    next_serial = int(last_serial) + 1
                except Exception:
                    # fallback if last value isn't int
                    # scan upward for last numeric serial
                    next_serial = 1
                    for r in range(ws.max_row, 0, -1):
                        val = ws.cell(row=r, column=1).value
                        try:
                            if val is not None:
                                next_serial = int(val) + 1
                                break
                        except Exception:
                            continue
                serial_to_write = next_serial
            else:
                serial_to_write = 1
        except Exception:
            serial_to_write = 1

        # Append attendance row (use student ID and name from recognized data)
        try:
            # attendance holds [serial_no_from_students_file, student_id, student_name, date, time]
            _, student_id, student_name, date_val, time_val = attendance
            ws.append([serial_to_write, student_id, student_name, date_val, time_val])
            wb.save(attendance_file)
            autofit_excel(attendance_file)
        except Exception as e:
            mess._show(title='Error', message=f'Failed to save attendance: {e}')
            try:
                wb.save(attendance_file)
            except Exception:
                pass
            cam.release()
            cv2.destroyAllWindows()
            return

        # Load and show attendance in Treeview (latest first)
        try:
            df_att = pd.read_excel(attendance_file)
            # Clear treeview then insert rows (we cleared earlier)
            for _, row in df_att.iterrows():
                try:
                    if pd.notna(row['S.No']) or pd.notna(row['S.No'.upper()]):
                        # handle both header cases if necessary
                        s = row.get('S.No', row.get('S.NO', row.get('S.No.', None)))
                    # safe insertion
                    tv.insert('', 'end', text=row.iloc[0], values=(row.iloc[1], row.iloc[2], row.iloc[3], row.iloc[4]))
                except Exception:
                    vals = list(row)
                    if len(vals) >= 5:
                        tv.insert('', 'end', text=vals[0], values=(vals[1], vals[2], vals[3], vals[4]))
        except Exception:
            pass

    cam.release()
    cv2.destroyAllWindows()

######################################## USED STUFFS ############################################

global key
key = ''

ts = time.time()
date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
day, month, year = date.split("-")

mont = {'01': 'January',
        '02': 'February',
        '03': 'March',
        '04': 'April',
        '05': 'May',
        '06': 'June',
        '07': 'July',
        '08': 'August',
        '09': 'September',
        '10': 'October',
        '11': 'November',
        '12': 'December'
        }

######################################## GUI FRONT-END ###########################################

# --- THEME AND STYLING ---
BG_COLOR = "#2d3436"
FG_COLOR = "#dfe6e9"
FRAME_COLOR = "#636e72"
ACCENT_COLOR = "#0984e3"
WARNING_COLOR = "#d63031"
FONT_FAMILY = "Helvetica"

window = tk.Tk()
window.geometry("1280x720")
window.title("Face Recognition Attendance System")
window.configure(background=BG_COLOR)
window.resizable(True, True)

# --- TTK STYLING ---
style = ttk.Style(window)
style.theme_use("clam")

# General widget styling
style.configure("TFrame", background=FRAME_COLOR)
style.configure("TLabel", background=FRAME_COLOR, foreground=FG_COLOR, font=(FONT_FAMILY, 12))
style.configure("Header.TLabel", font=(FONT_FAMILY, 16, "bold"))
style.configure("Title.TLabel", background=BG_COLOR, foreground=FG_COLOR, font=(FONT_FAMILY, 24, "bold"))
style.configure("Time.TLabel", background=BG_COLOR, foreground=ACCENT_COLOR, font=(FONT_FAMILY, 14, "bold"))
style.configure("Warning.TLabel", background=FRAME_COLOR, foreground=WARNING_COLOR, font=(FONT_FAMILY, 9))

style.configure("TEntry", fieldbackground="#b2bec3", foreground="#2d3436", font=(FONT_FAMILY, 12))
style.map("TEntry", foreground=[('focus', '#2d3436')])

style.configure("TButton", font=(FONT_FAMILY, 12), padding=10)
style.map("TButton",
          background=[('!active', '#b2bec3'), ('active', '#dfe6e9')],
          foreground=[('!active', '#2d3436'), ('active', '#2d3436')])

style.configure("Accent.TButton", font=(FONT_FAMILY, 12, "bold"), padding=10)
style.map("Accent.TButton",
          background=[('!active', ACCENT_COLOR), ('active', '#74b9ff')],
          foreground=[('!active', FG_COLOR), ('active', FG_COLOR)])

style.configure("Danger.TButton", font=(FONT_FAMILY, 10), padding=5)
style.map("Danger.TButton",
          background=[('!active', WARNING_COLOR), ('active', '#ff7675')],
          foreground=[('!active', FG_COLOR), ('active', FG_COLOR)])

# Treeview styling
style.configure("Treeview",
                background="#b2bec3",
                foreground="#2d3436",
                fieldbackground="#b2bec3",
                rowheight=25,
                font=(FONT_FAMILY, 11))
style.map("Treeview", background=[('selected', ACCENT_COLOR)])
style.configure("Treeview.Heading", font=(FONT_FAMILY, 12, "bold"), padding=5)

# --- HEADER ---
header_frame = ttk.Frame(window, style="TFrame", padding=(20, 10))
header_frame.pack(fill="x", side="top")
header_frame.configure(style="BG.TFrame")

message3 = ttk.Label(header_frame, text="Face Recognition Based Attendance Monitoring System", style="Title.TLabel")
message3.pack(side="left", expand=True)

time_frame = ttk.Frame(header_frame, style="BG.TFrame")
time_frame.pack(side="right")

datef = ttk.Label(time_frame, text=f"{day}-{mont[month]}-{year}", style="Time.TLabel")
datef.pack(side="top")

clock = ttk.Label(time_frame, style="Time.TLabel")
clock.pack(side="top")
tick()

# --- MAIN CONTENT ---
main_container = ttk.Frame(window, padding=20)
main_container.pack(fill="both", expand=True)
main_container.columnconfigure(0, weight=1, uniform="group1")
main_container.columnconfigure(1, weight=1, uniform="group1")
main_container.rowconfigure(0, weight=1)

# --- LEFT FRAME (FRAME 1 - Attendance) ---
frame1 = ttk.Frame(main_container, padding=20)
frame1.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
frame1.columnconfigure(0, weight=1)
frame1.rowconfigure(2, weight=1) # Treeview row

head1 = ttk.Label(frame1, text="Attendance Records", style="Header.TLabel", anchor="center")
head1.grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="ew")

# --- Email Section ---
email_frame = ttk.LabelFrame(frame1, text="Email Attendance Report", padding=15)
email_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 15))
email_frame.columnconfigure(1, weight=1)

ttk.Label(email_frame, text="Recipient:").grid(row=0, column=0, sticky="w", padx=(0, 10))
recipient_email_entry = ttk.Entry(email_frame, width=25)
recipient_email_entry.grid(row=0, column=1, sticky="ew")

email_domains = ["gmail.com", "yahoo.com", "hotmail.com"]
domain_var = tk.StringVar(frame1)
domain_var.set(email_domains[0])
domain_dropdown = ttk.OptionMenu(email_frame, domain_var, email_domains[0], *email_domains)
domain_dropdown.grid(row=0, column=2, padx=(5, 10))

send_email_button = ttk.Button(email_frame, text="Send", command=lambda: send_email(), width=8)
send_email_button.grid(row=0, column=3)

# --- Treeview ---
tree_frame = ttk.Frame(frame1)
tree_frame.grid(row=2, column=0, columnspan=3, sticky="nsew", pady=(0, 15))
tree_frame.columnconfigure(0, weight=1)
tree_frame.rowconfigure(0, weight=1)

tv = ttk.Treeview(tree_frame, columns=('id', 'name', 'date', 'time'))
tv.grid(row=0, column=0, sticky="nsew")

tv.column('#0', width=60, anchor='center')
tv.column('id', width=80, anchor='center')
tv.column('name', width=130, anchor='center')
tv.column('date', width=130, anchor='center')
tv.column('time', width=130, anchor='center')

tv.heading('#0', text='S.No.')
tv.heading('id', text='ID')
tv.heading('name', text='Name')
tv.heading('date', text='Date')
tv.heading('time', text='Time')

scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=tv.yview)
scroll.grid(row=0, column=1, sticky='ns')
tv.configure(yscrollcommand=scroll.set)

scroll_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=tv.xview)
scroll_x.grid(row=1, column=0, sticky='ew')
tv.configure(xscrollcommand=scroll_x.set)

# --- Action Buttons (Left) ---
action_frame1 = ttk.Frame(frame1)
action_frame1.grid(row=3, column=0, columnspan=3, sticky="ew")
action_frame1.columnconfigure(0, weight=1)

trackImg = ttk.Button(action_frame1, text="Take Attendance", command=TrackImages, style="Accent.TButton")
trackImg.grid(row=0, column=0, sticky="ew", pady=(0, 10))

quitWindow = ttk.Button(action_frame1, text="Quit", command=window.destroy)
quitWindow.grid(row=1, column=0, sticky="ew")


# --- RIGHT FRAME (FRAME 2 - Registration) ---
frame2 = ttk.Frame(main_container, padding=20)
frame2.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
frame2.columnconfigure(0, weight=1)

head2 = ttk.Label(frame2, text="New Registration", style="Header.TLabel", anchor="center")
head2.grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="ew")

# --- Input Fields ---
ttk.Label(frame2, text="Enter ID").grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 2))
txt = ttk.Entry(frame2, width=40)
txt.grid(row=2, column=0, columnspan=2, sticky="ew")
clearButton = ttk.Button(frame2, text="Clear", command=clear, width=8)
clearButton.grid(row=2, column=2, padx=(10, 0))
txt_warning = ttk.Label(frame2, text="", style="Warning.TLabel")
txt_warning.grid(row=3, column=0, columnspan=3, sticky="w", pady=(0, 10))
check_capslock_continuous(txt, txt_warning)

ttk.Label(frame2, text="Enter Name").grid(row=4, column=0, columnspan=3, sticky="w", pady=(10, 2))
txt2 = ttk.Entry(frame2, width=40)
txt2.grid(row=5, column=0, columnspan=2, sticky="ew")
clearButton2 = ttk.Button(frame2, text="Clear", command=clear2, width=8)
clearButton2.grid(row=5, column=2, padx=(10, 0))
txt2_warning = ttk.Label(frame2, text="", style="Warning.TLabel")
txt2_warning.grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 10))
check_capslock_continuous(txt2, txt2_warning)

# --- Action Buttons (Right) ---
takeImg = ttk.Button(frame2, text="1. Take Images", command=TakeImages)
takeImg.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(20, 10))

trainImg = ttk.Button(frame2, text="2. Save Profile", command=psw, style="Accent.TButton")
trainImg.grid(row=8, column=0, columnspan=3, sticky="ew")

# --- Status Messages ---
status_frame = ttk.Frame(frame2, padding=(0, 20))
status_frame.grid(row=9, column=0, columnspan=3, sticky="sew", pady=(20, 0))
frame2.rowconfigure(9, weight=1)
status_frame.columnconfigure(0, weight=1)

message1 = ttk.Label(status_frame, text="1) Take Images  >>>  2) Save Profile", anchor="center")
message1.grid(row=0, column=0, sticky="ew", pady=5)

message = ttk.Label(status_frame, text="", anchor="center", font=(FONT_FAMILY, 12, "bold"))
message.grid(row=1, column=0, sticky="ew", pady=5)

# --- Danger Zone ---
danger_zone = ttk.LabelFrame(frame2, text="Danger Zone", padding=15)
danger_zone.grid(row=10, column=0, columnspan=3, sticky="ew", pady=(20, 0))
danger_zone.columnconfigure(0, weight=1)
danger_zone.columnconfigure(1, weight=1)

delete_registration_button = ttk.Button(danger_zone, text="Delete Registrations", command=lambda: delete_registration_xlsx(), style="Danger.TButton")
delete_registration_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))

delete_attendance_button = ttk.Button(danger_zone, text="Delete Attendance", command=lambda: delete_attendance_xlsx(), style="Danger.TButton")
delete_attendance_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))

delete_images_button = ttk.Button(danger_zone, text="Delete Images", command=lambda: delete_registered_images(), style="Danger.TButton")
delete_images_button.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))

# Ensure StudentDetails file exists and compute initial count
if not os.path.exists(student_details_file):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(['SERIAL NO.', 'ID', 'NAME'])
        wb.save(student_details_file)
        autofit_excel(student_details_file)
    except Exception:
        pass

res = compute_registration_count()
message.configure(text='Total Registrations: ' + str(res))

##################### MENUBAR #################################

menubar = tk.Menu(window)
filemenu = tk.Menu(menubar, tearoff=0)
filemenu.add_command(label='Change Password', command=change_pass)
filemenu.add_command(label='Contact Us', command=contact)
filemenu.add_separator()
filemenu.add_command(label='Exit', command=window.destroy)
menubar.add_cascade(label='Help', menu=filemenu)

###################### Email Sending Logic ##################################
def send_email():
    recipient_email = recipient_email_entry.get().strip()
    selected_domain = domain_var.get().strip()

    if not recipient_email:
        mess._show(title='Error', message='Please enter a recipient email address.')
        return

    if '@' not in recipient_email:
        recipient_email = recipient_email + "@" + selected_domain

    # Hardcoded sender credentials
    from_email = "kapkotiprabhat@gmail.com"
    password = "eooj lynz auoi jusy"

    attendance_file = os.path.join("Attendance", "Attendance.xlsx")
    if not os.path.isfile(attendance_file):
        mess._show(title='Error', message=f'Attendance file not found: {attendance_file}')
        return

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = recipient_email
        msg['Subject'] = "Attendance Report"

        body = "Please find attached the attendance report."
        msg.attach(MIMEText(body, 'plain'))

        with open(attendance_file, "rb") as attachment:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename={os.path.basename(attendance_file)}")
            msg.attach(part)

        # Use a generic SMTP server logic, assuming Gmail for this example
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, recipient_email, msg.as_string())
        server.quit()
        mess._show(title='Success', message='Attendance report sent successfully.')
    except Exception as e:
        mess._show(title='Error', message=f'Failed to send email. Check credentials and SMTP settings.\n({e})')

###################### Delete functions ##################################
def delete_registration_xlsx():
    if mess.askyesno("Confirm", "Are you sure you want to delete all student registration data? This cannot be undone."):
        registration_xlsx_path = student_details_file
        if os.path.exists(registration_xlsx_path):
            try:
                os.remove(registration_xlsx_path)
                mess.showinfo("Success", "Registration XLSX file deleted successfully.")
                # Recreate empty registration file
                wb = Workbook()
                ws = wb.active
                ws.title = "Students"
                ws.append(['SERIAL NO.', 'ID', 'NAME'])
                wb.save(registration_xlsx_path)
                autofit_excel(registration_xlsx_path)
                update_registration_counter()
            except Exception as e:
                mess.showinfo("Error", f"Failed to delete Registration XLSX: {e}")
        else:
            mess.showinfo("Info", "Registration XLSX file not found.")

def delete_attendance_xlsx():
    if mess.askyesno("Confirm", "Are you sure you want to delete the attendance history? This cannot be undone."):
        attendance_xlsx_path = os.path.join("Attendance", "Attendance.xlsx")
        if os.path.exists(attendance_xlsx_path):
            try:
                os.remove(attendance_xlsx_path)
                mess.showinfo("Success", f"Attendance XLSX file deleted successfully.")
                # Clear the treeview
                for k in tv.get_children():
                    tv.delete(k)
            except Exception as e:
                mess.showinfo("Error", f"Failed to delete attendance: {e}")
        else:
            mess.showinfo("Info", f"Attendance XLSX file not found.")

def delete_registered_images():
    if mess.askyesno("Confirm", "Are you sure you want to delete all registered face images? This will require re-training all profiles."):
        folder_path = "TrainingImage/"
        if os.path.exists(folder_path):
            try:
                shutil.rmtree(folder_path)
                os.makedirs(folder_path) # Recreate the folder
                mess.showinfo("Success", "Registered images deleted successfully.")
            except Exception as e:
                mess.showinfo("Error", f"Failed to delete registered images: {e}")
        else:
            mess.showinfo("Info", "TrainingImage folder not found.")

##################### END ######################################
window.configure(menu=menubar)
window.mainloop()
